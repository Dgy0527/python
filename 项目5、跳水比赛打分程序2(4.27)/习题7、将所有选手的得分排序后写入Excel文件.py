#将所有选手的得分排序后写入Excel文件
import random 
from openpyxl import load_workbook,Workbook

wb=load_workbook('选手信息.xlsx')
sheetname=wb.sheetnames[0]
ws=wb[sheetname]

dic={}

for rn in range(1,ws.max_row+1):
    temp_list=[]
    key=ws['A'+str(rn)].value
    w1=ws['B'+str(rn)].value
    w2=ws['C'+str(rn)].value
    w3=ws['D'+str(rn)].value

    temp_list=[w1,w2,w3]
    dic[key]=temp_list

num_list=[] #一个选手你的分数
result_list=[] #所有选手的分数
aver_list=[]
scor_list=[]

rows=8 #8个选手
columns=10 #10个裁判

for i in range(rows):
    num_list.append([])
    result_list.append([])

    for j in range(columns):
        num=random.randint(0,10) #随机生成0-10之间的一个整数
        num_list[i].append(num) #添加到选手列表中

    num_list[i].sort() #对列表进行升序排序

    del num_list[i][0] #删除第一个元素
    del num_list[i][-1] #删除最后一个元素

    aver=round(sum(num_list[i])/len(num_list[i]),2) #求列表的平均值
    scor=round(aver*dic['s0'+str(i+1)][2],2) #求最后得分

    result_list[i].append('s0'+str(i+1))

    for n in range(3):
        result_list[i].append(dic['s0'+str(i+1)][n])

    result_list[i].append(aver)
    result_list[i].append(scor)

last_result=sorted(result_list,key=lambda e:e[5],reverse=True)

for n in range(rows):
    last_result[n].append(n+1)

for i in last_result:
    print(i)

#开始写入代码
wb=Workbook()
sheet=wb.active
sheet.title='成绩表'
sheet_field=['选手ID','姓名','国家','难度系数','平均分','最后得分','名次']
sheet.append(sheet_field)

for a in last_result:
    sheet.append(a)
wb.save('跳水比赛成绩登记表.xlsx')



'''
对上面代码的改进:

import random
import os
from openpyxl import load_workbook, Workbook

# -------------------- 1. 从 Excel 读取选手信息（跳过表头） --------------------
try:
    wb_info = load_workbook('选手信息.xlsx')
    ws = wb_info[wb_info.sheetnames[0]]
except FileNotFoundError:
    print("错误：找不到文件 '选手信息.xlsx'，请检查路径！")
    exit()

players = {}  # 选手信息字典 {ID: {'姓名':..., '国家':..., '难度系数':...}}

# 假设第1行是表头,数据从第2行开始
for row in range(2, ws.max_row + 1):
    pid = ws[f'A{row}'].value
    if pid is None:          # 跳过空行
        continue
    
    name = ws[f'B{row}'].value or ''
    country = ws[f'C{row}'].value or ''
    diff_raw = ws[f'D{row}'].value
    
    # 将难度系数转为浮点数,无效时设为默认值3.0
    try:
        difficulty = float(diff_raw) if diff_raw is not None else 3.0
    except (TypeError, ValueError):
        print(f"警告：选手 {pid} 的难度系数无效，已设为 3.0")
        difficulty = 3.0
    
    players[pid] = {
        '姓名': name,
        '国家': country,
        '难度系数': difficulty
    }

if not players:
    print("错误：未读取到任何选手信息，程序终止。")
    exit()

# -------------------- 2. 模拟裁判打分并计算最终成绩 --------------------
JUDGE_COUNT = 10  # 裁判人数
result_list = []  # 每个元素：[ID, 姓名, 国家, 难度系数, 平均分, 最后得分]

for pid, info in players.items():
    # 10名裁判随机打分(0~10分)
    raw_scores = [random.randint(0, 10) for _ in range(JUDGE_COUNT)]
    
    # 排序后去掉一个最高分和一个最低分（保留原始数据）
    sorted_scores = sorted(raw_scores)
    valid_scores = sorted_scores[1:-1]   # 去掉首尾元素
    
    # 计算平均分（保留两位小数）
    avg_score = round(sum(valid_scores) / len(valid_scores), 2)
    # 最终得分 = 平均分 * 难度系数
    final_score = round(avg_score * info['难度系数'], 2)
    
    result_list.append([
        pid,
        info['姓名'],
        info['国家'],
        info['难度系数'],
        avg_score,
        final_score
    ])

# -------------------- 3. 按最终得分降序排列，并分配名次（处理并列） --------------------
result_list.sort(key=lambda x: x[5], reverse=True)  # 按第6列(最终得分)降序

rank = 1
prev_score = None
for i, row in enumerate(result_list):
    current_score = row[5]
    if prev_score is not None and current_score != prev_score:
        rank = i + 1        # 得分不同时，名次为当前位次
    row.append(rank)        # 添加名次列
    prev_score = current_score

# 预览结果
for item in result_list:
    print(item)

# -------------------- 4. 将结果写入新的 Excel 文件 --------------------
wb_out = Workbook()
sheet = wb_out.active
sheet.title = '成绩表'

# 写入表头
headers = ['选手ID', '姓名', '国家', '难度系数', '平均分', '最后得分', '名次']
sheet.append(headers)

# 逐行写入数据
for row in result_list:
    sheet.append(row)

# 保存到当前脚本所在目录
script_dir = os.path.dirname(os.path.abspath(__file__))
save_path = os.path.join(script_dir, '跳水比赛成绩登记表.xlsx')
wb_out.save(save_path)

print(f"成绩表已生成：{save_path}")

'''