#将所有选手的得分导入二位列表
import random
from openpyxl import load_workbook

wb=load_workbook('选手信息.xlsx')
sheetname=wb.sheetnames[0]
ws=wb[sheetname]

dic={}

for rn in range(1,ws.max_row+1):
    temp_list=[]
    key=ws['A'+str(rn)].value
    w1=ws['B'+str(rn)].value
    w2=ws['C'+str(rn)].value
    w3=ws['D'+str(rn)].value

    temp_list=[w1,w2,w3]
    dic[key]=temp_list

num_list=[] #一个选手的分数
result_list=[] #所有选手的分数
aver_list=[]
scor_list=[]

rows=8 #8个选手
columns=10 #10个裁判

for i in range(rows):
    num_list.append([])
    result_list.append([])

    for j in range(columns):
        num=random.randint(0,10) #随机生成0-10之间的一个整数
        num_list[i].append(num) #添加到选手列表中

    num_list[i].sort() #对列表进行升序排序

    del num_list[i][0] #删除第一个元素
    del num_list[i][-1] #删除最后一个元素
    
    aver=round(sum(num_list[i])/len(num_list[i]),2) #求列表的平均值
    scor=round(aver*dic['s0'+str(i+1)][2],2) #求最后得分

    result_list[i].append('s0'+str(i+1))

    for n in range(3):
        result_list[i].append(dic['s0'+str(i+1)][n])
    result_list[i].append(aver)
    result_list[i].append(scor)
    print(result_list[i])




'''
对上面代码的改进:

import random
from openpyxl import load_workbook

# 1. 读取选手信息（自动适应行数）
try:
    wb = load_workbook('选手信息.xlsx')
    sheet = wb[wb.sheetnames[0]]
except FileNotFoundError:
    print("错误：找不到文件 '选手信息.xlsx'")
    exit()

players = {}  # 选手信息:ID -> {姓名, 国家, 难度系数}
for row in range(1, sheet.max_row + 1):
    pid = sheet[f'A{row}'].value
    if pid is None:
        continue  # 跳过空行
    name = sheet[f'B{row}'].value or ''
    country = sheet[f'C{row}'].value or ''
    diff = sheet[f'D{row}'].value
    try:
        diff = float(diff)
    except (TypeError, ValueError):
        print(f"警告：选手 {pid} 的难度系数无效，使用默认值 3.0")
        diff = 3.0
    players[pid] = {
        '姓名': name,
        '国家': country,
        '难度系数': diff
    }

JUDGE_COUNT = 10

result_list = []   # 最终结果二维列表

# 2. 按字典中的实际选手循环处理
for pid, info in players.items():
    # 生成裁判分数
    scores = [random.randint(0, 10) for _ in range(JUDGE_COUNT)]
    # 去掉一个最高分和一个最低分（不破坏原列表）
    sorted_scores = sorted(scores)
    valid_scores = sorted_scores[1:-1]

    avg_score = round(sum(valid_scores) / len(valid_scores), 2)
    final_score = round(avg_score * info['难度系数'], 2)

    # 组合结果
    row_data = [
        pid,
        info['姓名'],
        info['国家'],
        info['难度系数'],
        avg_score,
        final_score
    ]
    result_list.append(row_data)

# 3. 输出查看
for row in result_list:
    print(row)

print(f"\n共处理 {len(result_list)} 位选手，结果已存入 result_list 二维列表。")

'''

