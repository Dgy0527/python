from openpyxl import load_workbook #导入读子模块

wb=load_workbook('选手信息.xlsx') #1、将选手信息打开并赋值给wb
sheet_name=wb.sheetnames[0] #2、获取wb工作簿的第一个工作表
ws=wb[sheet_name] #3、读取指定的工作表sheet_name
dic={} #定义一个新字典

for rn in range(1,ws.max_row+1):
    temp_list=[]
    key=ws['A'+str(rn)].value
    w1=ws['B'+str(rn)].value
    w2=ws['C'+str(rn)].value
    w3=ws['D'+str(rn)].value
    temp_list=[w1,w2,w3]
    dic[key]=temp_list
print(dic)

for dic_id,info in dic.items():
    print(f"{dic_id}{info}")


'''
上面代码存在的问题:1、不能去掉xlsx里面的表头
对上面代码的改进:

import os
from openpyxl import load_workbook

# 1. 明确指定路径（如果文件在你上一次写入的那个文件夹，请把下面这行改成绝对路径）
# 例如:file_path = r'D:\Git\项目4、跳水比赛打分程序1(4.20)\选手信息.xlsx'
file_path = '选手信息.xlsx' 

if not os.path.exists(file_path):
    print(f"错误：找不到文件 {file_path}，请检查文件是否在当前目录下！")
else:
    try:
        wb = load_workbook(file_path)
        ws = wb.active  # 直接获取当前激活的工作表，更简洁
        dic = {}

        # 2. 使用 iter_rows 并从第 2 行开始读取(min_row=2),自动跳过表头
        # values_only=True 表示直接获取单元格的值，省去 .value 的写法
        for row in ws.iter_rows(min_row=2, max_col=4, values_only=True):
            key = row[0]  # A列(选手ID)
            if key is not None:  # 过滤掉可能存在的空行
                # B、C、D列分别对应:姓名、国籍、难度系数
                temp_list = [row[1], row[2], row[3]]
                dic[key] = temp_list

        # 3. 输出结果
        print(dic)
        print("--- 选手详细信息 ---")
        for dic_id, info in dic.items():
            print(f"{dic_id} {info}")

    except PermissionError:
        print("错误：文件被 Excel 或其他程序占用，请关闭文件后再运行！")
    except Exception as e:
        print(f"发生未知错误：{e}")
'''
