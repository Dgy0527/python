#向Excel文件中写入数据
#库openpyxl:用于Excel文件进行操作的模块
from openpyxl import Workbook #导入子模块

'''
wb=Workbook() #1、创建一个工作簿wb
sheet=wb.active #2、创建并激活一个工作表sheet
sheet.title='销售表' #3、修改工作表的标签
sheet.append(['选手ID','姓名','国籍','难度系数']) #4、向单元格中添加数据
sheet.append(['s01','马丁','意大利',3.1])
wb.save('save.xlsx') #5、保存工作簿
'''

list=[
    ['选手ID','姓名','国籍','难度系数'],
    ['s02','马丁','意大利',3.1],
    ['s03','托马斯','西班牙',2.8],
    ['s04','理查德','意大利',2.5],
    ['s05','马丁','意大利',3.1],
    ['s06','马丁','意大利',3.1],
    ['s07','马丁','意大利',3.1],
    ['s08','马丁','意大利',3.1],
]

wb=Workbook()
sheet=wb.active
sheet.title='选手信息'
for row in list:
    sheet.append(row)
wb.save('选手信息表.xlsx')


'''
对上面代码的改进:

from openpyxl import Workbook
import os

# 数据列表（避免使用 list 作为变量名）
data = [
    ['选手ID', '姓名', '国籍', '难度系数'],
    ['s02', '马丁', '意大利', 3.1],
    ['s03', '托马斯', '西班牙', 2.8],
    ['s04', '理查德', '意大利', 2.5],
    ['s05', '马丁', '意大利', 3.1],
    ['s06', '马丁', '意大利', 3.1],
    ['s07', '马丁', '意大利', 3.1],
    ['s08', '马丁', '意大利', 3.1],
]

file_name = '选手信息表.xlsx'

# 检查文件是否已存在（非必须，但提示更友好）
if os.path.exists(file_name):
    print(f"提示：文件 {file_name} 已存在，即将覆盖！")

# 使用 with 上下文管理器，保存后自动关闭文件资源
try:
    with Workbook() as wb:
        sheet = wb.active
        sheet.title = '选手信息'
        
        # 逐行写入数据
        for row in data:
            sheet.append(row)
        
        # 保存文件
        wb.save(file_name)
        print(f"数据成功写入 {file_name}")

except PermissionError:
    print(f"错误：请关闭已打开的 {file_name} 文件，然后再运行程序！")
except Exception as e:
    print(f"保存失败，错误信息：{e}")

'''